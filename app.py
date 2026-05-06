from flask import Flask, request, render_template, send_file, jsonify
from werkzeug.utils import secure_filename
import os
import tempfile
import zipfile
from io import BytesIO
import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extraer_datos_pdf(pdf_path):
    """Extrae información del PDF de nota de crédito"""
    with pdfplumber.open(pdf_path) as pdf:
        texto_completo = ""
        for page in pdf.pages:
            texto_completo += page.extract_text() + "\n"
    
    # Extraer número de documento
    match_doc = re.search(r'NF(\d+)', texto_completo)
    num_doc = f"NF{match_doc.group(1)}" if match_doc else "DESCONOCIDO"
    
    # Extraer fecha
    match_fecha = re.search(r'(\d{2}/\w{3}/\d{4}\s+\d{2}:\d{2})', texto_completo)
    fecha = match_fecha.group(1) if match_fecha else "DESCONOCIDO"
    
    # Extraer pedido
    match_pedido = re.search(r'G(\d+)', texto_completo)
    pedido = f"G{match_pedido.group(1)}" if match_pedido else "DESCONOCIDO"
    
    # Extraer productos (buscar patrones de productos)
    productos = []
    
    # Patrón para líneas de productos
    # Formato: CODIGO CANTIDAD CVE_PROD UNIDAD CLAVE DESCRIPCION BARCODE PRECIO IMPORTE
    patron = r'(\d+)\s+1\s+26111707\s+PZ\s+H87\s+([\w\s\-]+?)\s+(75\d+)?\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})'
    
    for match in re.finditer(patron, texto_completo):
        codigo = match.group(1)
        descripcion = match.group(2).strip()
        barcode = match.group(3) if match.group(3) else ""
        precio = float(match.group(4).replace(',', ''))
        importe = float(match.group(5).replace(',', ''))
        cantidad = 1
        
        productos.append([codigo, descripcion, barcode, 'PZ', precio, cantidad])
    
    return {
        'documento': num_doc,
        'fecha': fecha,
        'pedido': pedido,
        'productos': productos
    }

def consolidar_productos(productos_list):
    """Consolida productos por descripción"""
    consolidado = defaultdict(lambda: {
        'codigo': '', 'barras': '', 'unidad': '', 
        'cantidad': 0, 'precios': [], 'importe_total': 0
    })
    
    for producto in productos_list:
        codigo, desc, barras, unidad, precio, cant = producto
        key = desc
        if not consolidado[key]['codigo']:
            consolidado[key]['codigo'] = codigo
            consolidado[key]['barras'] = barras
            consolidado[key]['unidad'] = unidad
        consolidado[key]['cantidad'] += cant
        consolidado[key]['precios'].append(precio)
        consolidado[key]['importe_total'] += precio * cant
    
    for key in consolidado:
        precios = consolidado[key]['precios']
        consolidado[key]['precio_prom'] = sum(precios) / len(precios)
    
    return consolidado

def crear_excel_consolidado(datos_doc, filepath):
    """Crea archivo Excel consolidado"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Nota de Crédito"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Configurar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Título
    ws['A1'] = 'NOTA DE CRÉDITO (CONSOLIDADO)'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Información del documento
    row = 3
    ws[f'A{row}'] = 'Documento:'
    ws[f'B{row}'] = datos_doc['documento']
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Fecha:'
    ws[f'B{row}'] = datos_doc['fecha']
    
    row += 1
    ws[f'A{row}'] = 'Pedido Cliente:'
    ws[f'B{row}'] = datos_doc['pedido']
    
    # Headers de empresa
    row += 2
    ws[f'A{row}'] = 'EMISOR'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    ws[f'A{row}'] = 'JOHNSON CONTROLS ENTERPRISES MEXICO S DE R L DE C V'
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    ws[f'A{row}'] = 'RFC: JCA100604EF4'
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 2
    ws[f'A{row}'] = 'CLIENTE'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    ws[f'A{row}'] = 'ENERGY PARTS'
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    ws[f'A{row}'] = 'RFC: EPA191009JJ0'
    ws.merge_cells(f'A{row}:G{row}')
    
    # Headers de tabla
    row += 2
    header_row = row
    headers = ['CÓDIGO', 'CANTIDAD', 'DESCRIPCIÓN', 'CÓDIGO BARRAS', 'UNIDAD', 'P. PROMEDIO', 'IMPORTE TOTAL']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Datos consolidados
    consolidado = consolidar_productos(datos_doc['productos'])
    productos_ordenados = sorted(consolidado.items(), key=lambda x: x[0])
    
    row = header_row + 1
    primera_fila = row
    
    for descripcion, datos in productos_ordenados:
        ws.cell(row=row, column=1, value=datos['codigo'])
        ws.cell(row=row, column=2, value=datos['cantidad'])
        ws.cell(row=row, column=3, value=descripcion)
        ws.cell(row=row, column=4, value=datos['barras'])
        ws.cell(row=row, column=5, value=datos['unidad'])
        ws.cell(row=row, column=6, value=datos['precio_prom']).number_format = '$#,##0.00'
        ws.cell(row=row, column=7, value=datos['importe_total']).number_format = '$#,##0.00'
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border_thin
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='left' if col == 3 else 'center')
        
        row += 1
    
    ultima_fila = row - 1
    
    # Totales
    row += 1
    ws[f'F{row}'] = 'SUBTOTAL:'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'G{row}'] = f'=SUM(G{primera_fila}:G{ultima_fila})'
    ws[f'G{row}'].number_format = '$#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    fila_subtotal = row
    
    row += 1
    ws[f'F{row}'] = 'IVA 16%:'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'G{row}'] = f'=G{fila_subtotal}*0.16'
    ws[f'G{row}'].number_format = '$#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    fila_iva = row
    
    row += 1
    ws[f'F{row}'] = 'TOTAL:'
    ws[f'F{row}'].font = Font(bold=True, size=12)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'F{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws[f'G{row}'] = f'=G{fila_subtotal}+G{fila_iva}'
    ws[f'G{row}'].number_format = '$#,##0.00'
    ws[f'G{row}'].font = Font(bold=True, size=12)
    ws[f'G{row}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    row += 2
    num_original = len(datos_doc['productos'])
    num_consolidado = len(consolidado)
    ws[f'A{row}'] = f'Nota: Este documento consolida {num_original} líneas en {num_consolidado} productos únicos'
    ws[f'A{row}'].font = Font(italic=True, size=9)
    ws.merge_cells(f'A{row}:G{row}')
    
    wb.save(filepath)

@app.route('/')
def index():
    return render_template('index.html')

def crear_excel_consolidado_general(todos_los_datos, filepath):
    """Crea archivo Excel consolidado GENERAL de todos los PDFs"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado General"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Configurar anchos
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Título
    ws['A1'] = f'CONSOLIDADO GENERAL - {len(todos_los_datos)} DOCUMENTOS'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    ws[f'A{row}'] = 'Documentos:'
    docs = ', '.join([d['documento'] for d in todos_los_datos])
    ws[f'B{row}'] = docs
    ws[f'B{row}'].font = Font(bold=True)
    ws.merge_cells(f'B{row}:G{row}')
    
    row += 1
    ws[f'A{row}'] = 'Cliente:'
    ws[f'B{row}'] = 'ENERGY PARTS (EPA191009JJ0)'
    
    row += 1
    ws[f'A{row}'] = 'Emisor:'
    ws[f'B{row}'] = 'JOHNSON CONTROLS ENTERPRISES MEXICO (JCA100604EF4)'
    
    # Consolidar TODOS los productos
    todos_productos = []
    for datos_doc in todos_los_datos:
        todos_productos.extend(datos_doc['productos'])
    
    consolidado = consolidar_productos(todos_productos)
    
    # Headers de tabla
    row += 2
    header_row = row
    headers = ['CÓDIGO', 'DESCRIPCIÓN', 'CÓDIGO BARRAS', 'UNIDAD', 'CANTIDAD', 'P. PROMEDIO', 'IMPORTE TOTAL']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin
    
    # Datos consolidados
    productos_ordenados = sorted(consolidado.items(), key=lambda x: x[0])
    
    row = header_row + 1
    primera_fila = row
    
    for descripcion, datos in productos_ordenados:
        ws.cell(row=row, column=1, value=datos['codigo'])
        ws.cell(row=row, column=2, value=descripcion)
        ws.cell(row=row, column=3, value=datos['barras'])
        ws.cell(row=row, column=4, value=datos['unidad'])
        ws.cell(row=row, column=5, value=datos['cantidad'])
        ws.cell(row=row, column=6, value=datos['precio_prom']).number_format = '$#,##0.00'
        ws.cell(row=row, column=7, value=datos['importe_total']).number_format = '$#,##0.00'
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border_thin
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='left' if col == 2 else 'center')
        
        row += 1
    
    ultima_fila = row - 1
    
    # TOTALES
    row += 1
    ws[f'E{row}'] = 'TOTAL PIEZAS:'
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'E{row}'].alignment = Alignment(horizontal='right')
    ws[f'F{row}'] = f'=SUM(E{primera_fila}:E{ultima_fila})'
    ws[f'F{row}'].font = Font(bold=True, size=12)
    ws[f'F{row}'].alignment = Alignment(horizontal='center')
    ws[f'F{row}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    row += 1
    ws[f'F{row}'] = 'SUBTOTAL:'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'G{row}'] = f'=SUM(G{primera_fila}:G{ultima_fila})'
    ws[f'G{row}'].number_format = '$#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    fila_subtotal = row
    
    row += 1
    ws[f'F{row}'] = 'IVA 16%:'
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'G{row}'] = f'=G{fila_subtotal}*0.16'
    ws[f'G{row}'].number_format = '$#,##0.00'
    ws[f'G{row}'].font = Font(bold=True)
    fila_iva = row
    
    row += 1
    ws[f'F{row}'] = 'TOTAL GENERAL:'
    ws[f'F{row}'].font = Font(bold=True, size=12)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    ws[f'F{row}'].fill = yellow_fill
    ws[f'G{row}'] = f'=G{fila_subtotal}+G{fila_iva}'
    ws[f'G{row}'].number_format = '$#,##0.00'
    ws[f'G{row}'].font = Font(bold=True, size=12)
    ws[f'G{row}'].fill = yellow_fill
    
    # RESUMEN POR DOCUMENTO
    row += 3
    ws[f'A{row}'] = 'RESUMEN POR DOCUMENTO'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ws.merge_cells(f'A{row}:G{row}')
    
    row += 1
    ws[f'A{row}'] = 'DOCUMENTO'
    ws[f'B{row}'] = 'FECHA'
    ws[f'C{row}'] = 'PEDIDO'
    ws[f'D{row}'] = 'LÍNEAS ORIG.'
    ws[f'E{row}'] = 'PRODUCTOS ÚNICOS'
    ws[f'F{row}'] = 'SUBTOTAL'
    ws[f'G{row}'] = 'TOTAL'
    
    for col in range(1, 8):
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        ws.cell(row=row, column=col).border = border_thin
    
    for datos_doc in todos_los_datos:
        row += 1
        consolidado_doc = consolidar_productos(datos_doc['productos'])
        subtotal = sum(d['importe_total'] for d in consolidado_doc.values())
        total = subtotal * 1.16
        
        ws[f'A{row}'] = datos_doc['documento']
        ws[f'B{row}'] = datos_doc['fecha']
        ws[f'C{row}'] = datos_doc['pedido']
        ws[f'D{row}'] = len(datos_doc['productos'])
        ws[f'E{row}'] = len(consolidado_doc)
        ws[f'F{row}'] = subtotal
        ws[f'F{row}'].number_format = '$#,##0.00'
        ws[f'G{row}'] = total
        ws[f'G{row}'].number_format = '$#,##0.00'
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border_thin
    
    wb.save(filepath)

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files[]' not in request.files:
        return jsonify({'error': 'No se enviaron archivos'}), 400
    
    files = request.files.getlist('files[]')
    
    if not files or files[0].filename == '':
        return jsonify({'error': 'No se seleccionaron archivos'}), 400
    
    # Validar archivos
    for file in files:
        if not allowed_file(file.filename):
            return jsonify({'error': f'Archivo no permitido: {file.filename}'}), 400
    
    try:
        # Crear directorio temporal
        temp_dir = tempfile.mkdtemp()
        resultados = []
        todos_los_datos = []
        
        # Procesar cada PDF
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                pdf_path = os.path.join(temp_dir, filename)
                file.save(pdf_path)
                
                # Extraer datos
                datos = extraer_datos_pdf(pdf_path)
                todos_los_datos.append(datos)
                
                # Crear Excel individual
                excel_filename = f"Consolidado_{datos['documento']}.xlsx"
                excel_path = os.path.join(temp_dir, excel_filename)
                crear_excel_consolidado(datos, excel_path)
                
                resultados.append({
                    'pdf': filename,
                    'excel': excel_filename,
                    'documento': datos['documento'],
                    'productos_original': len(datos['productos']),
                    'productos_consolidado': len(consolidar_productos(datos['productos']))
                })
        
        # Crear consolidado GENERAL si hay más de 1 PDF
        if len(todos_los_datos) > 1:
            excel_general = os.path.join(temp_dir, 'CONSOLIDADO_GENERAL.xlsx')
            crear_excel_consolidado_general(todos_los_datos, excel_general)
        
        # Crear ZIP con todos los Excel
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Agregar archivos individuales
            for resultado in resultados:
                excel_path = os.path.join(temp_dir, resultado['excel'])
                zip_file.write(excel_path, resultado['excel'])
            
            # Agregar consolidado general si existe
            if len(todos_los_datos) > 1:
                zip_file.write(excel_general, 'CONSOLIDADO_GENERAL.xlsx')
        
        zip_buffer.seek(0)
        
        # Limpiar archivos temporales
        import shutil
        shutil.rmtree(temp_dir)
        
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name='notas_credito_consolidadas.zip'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
